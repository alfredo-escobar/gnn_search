import csv
import numpy as np
from openpyxl import Workbook

def save_most_similar_pairs(similarity):

    # 2D indexes for the similarity matrix:
    similarity_idx = np.triu_indices(similarity.shape[0], 1)

    # 1D array with similarity scores of upper half of matrix
    similarity_top = similarity[similarity_idx]

    #most_similar_idx1d = np.argsort(similarity_top)[-1:-51:-1]
    most_similar_idx1d = np.argsort(similarity_top)[0:100]
    most_similar_idx = [similarity_idx[0][most_similar_idx1d], similarity_idx[1][most_similar_idx1d]]

    # Create a new Workbook
    workbook = Workbook()
    sheet = workbook.active

    # Read data from the CSV file
    with open("./catalogues/UNIQLO/train_search.csv", 'r', encoding='utf-8') as csv_file:
        reader = csv.reader(csv_file, delimiter=';')
        csv_data = list(reader)

    # Iterate over the list 'most_similar_idx' and extract values from 'csv_data'
    for i in range(len(most_similar_idx[0])):
        # Extract values based on the indices from 'a'
        row_index_1 = most_similar_idx[0][i]
        row_index_2 = most_similar_idx[1][i]

        # Get values from 'csv_data' and store them in the XLSX sheet
        for col in range(5):
            value_1 = csv_data[row_index_1][col+4]
            sheet.cell(row=i+1, column=col+1, value=value_1)
        
        for col in range(5):
            value_2 = csv_data[row_index_2][col+4]
            sheet.cell(row=i+1, column=col+7, value=value_2)
        
        sheet.cell(row=i+1, column=15, value=similarity[row_index_1][row_index_2])

    # Save the XLSX file
    workbook.save("most_similar_pairs.xlsx")


def save_sims_and_probs(all_sims_and_probs):
    # Create a new workbook
    workbook = Workbook()

    # Iterate over each key in the dictionary and save it as a separate sheet
    for sheet_name, iter_dict in all_sims_and_probs.items():
        # Create a new sheet with the given sheet_name
        sheet = workbook.create_sheet(title=sheet_name)

        # Write the data to the sheet
        for col_index, col_name in enumerate(iter_dict):

            sheet.cell(row = 1, column = col_index+1, value = col_name)

            for row_index, value in enumerate(iter_dict[col_name]):
                sheet.cell(row = row_index+2, column = col_index+1, value = value.item())

    # Remove the default first sheet created by openpyxl
    workbook.remove(workbook['Sheet'])

    # Save the workbook to a .xlsx file
    workbook.save('sims_and_probs.xlsx')
